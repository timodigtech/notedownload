import re
import openpyxl

# 定义正则表达式
pattern = r'([\u4e00-\u9fa5a-zA-Z\s]+)(\d+\.?\d*)'

# 定义Excel文件名和工作表名
filename = 'data.xlsx'
sheetname = 'Sheet1'

# 创建Excel工作簿
wb = openpyxl.Workbook()
ws = wb.active

# 写入表头
ws['A1'] = '姓名'
ws['B1'] = '数字'

# 定义行数
row_num = 2

# 解析字符串
text = '蔡柳新1459.81,曹利平973.22,陈国俊1459.8,陈钶370.8,陈灵华4379.4,陈鸣宇1459.8,陈文超3524.4,丁国平a2433,冯旭3777,梁锴648.9,梁岳龙2919.6,林辉486.6,吕杰敏7102.5,潘军海486.6,沈波1459.8,沈海刚7204.8,沈吉良6325.8,史亮3406.2,童一凡1459.8,汪勇486.6,王遂函2435.4,王先法973.2,王一帆1418.7,吴峥嵘4866,许斌1946.4,严加费4634.4,杨瑾2919.6,杨进3826.5,虞洪1459.8,张翔宇1297.8,郑雪咏3661.2,钟欣973.2,周良晶857.4,朱越锋1946.7； '
matches = re.findall(pattern, text)

# 遍历匹配结果并写入Excel
for match in matches:
    name = match[0]
    value = float(match[1])
    ws.cell(row=row_num, column=1, value=name)
    ws.cell(row=row_num, column=2, value=value)
    row_num += 1

# 保存Excel文件
wb.save(filename)